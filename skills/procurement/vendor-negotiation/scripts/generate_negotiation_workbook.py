#!/usr/bin/env python3
"""Generate a vendor-negotiation preparation workbook.

Builds a 3-sheet .xlsx demonstrating the core data operation behind the
vendor-negotiation skill:

  Market Intel -> Negotiation Plan -> Outcome Log

- Sheet 1 "Market Intel": 6 realistic sourcing scenarios with benchmarks.
- Sheet 2 "Negotiation Plan": target/walk-away, BATNA, terms, concessions,
  must-haves, and strategy per supplier.
- Sheet 3 "Outcome Log": agreed outcome with formulas for savings vs. target
  and vs. walk-away.

Run:  python generate_negotiation_workbook.py [output.xlsx]
Requires: openpyxl
"""
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="7030A0")
HEADER_FONT = Font(bold=True, color="FFFFFF")

INTEL_HEADERS = [
    "Supplier", "Product/Category", "Market Price", "Their Price",
    "Competitor Price", "Availability", "Lead Time", "Quality Score",
]

MARKET_INTEL = [
    ("Coastal Components Inc.", "Control boards", 590, 640, 615, "In stock", "3 wks", 93),
    ("Meridian Metals Co.", "Steel plate 10mm", 210, 225, 218, "Limited", "5 wks", 90),
    ("Evergreen Chemicals", "Industrial solvent", 132, 148, 140, "In stock", "2 wks", 85),
    ("Anchor Logistics LLC", "Freight lane (regional)", 8_400, 9_100, 8_750, "Available", "n/a", 80),
    ("Falcon Precision Tooling", "CNC end mills", 45, 52, 49, "In stock", "4 wks", 88),
    ("Brightline Packaging", "Corrugated boxes", 2.9, 3.4, 3.1, "In stock", "3 wks", 74),
]

PLAN_HEADERS = [
    "Supplier", "Target Price", "Walk-Away Price", "BATNA",
    "Key Terms (warranty, payment, SLA)", "Concessions Available",
    "Must-Haves", "Nice-to-Haves", "Strategy",
]

NEGOTIATION_PLAN = [
    ("Coastal Components Inc.", 600, 630,
     "Second-source from Falcon at 625",
     "24-mo warranty; Net 60; 98% OTD SLA",
     "Order volume commit; multi-year term",
     "24-mo warranty; Net 60", "Consigned stock", "Collaborative"),
    ("Meridian Metals Co.", 205, 222,
     "Regional mill quote at 218",
     "12-mo price hold; Net 45; MOQ waiver",
     "Prompt-pay discount; forecast sharing",
     "12-mo price hold", "Freight included", "Competitive"),
    ("Evergreen Chemicals", 128, 145,
     "Alternate blend from ChemCo at 141",
     "REACH docs; Net 30; safety data on file",
     "Longer term; larger drum sizes",
     "REACH compliance", "Vendor-managed inventory", "Collaborative"),
    ("Anchor Logistics LLC", 8_200, 8_900,
     "Spot market at 8,750",
     "On-time SLA 97%; fuel-surcharge cap; Net 45",
     "Lane volume commit; annual tender",
     "Fuel-surcharge cap", "Dedicated capacity", "Competitive"),
    ("Falcon Precision Tooling", 44, 50,
     "In-house grind option at 51",
     "12-mo warranty; Net 45; AS9100 certs",
     "Blanket PO; tooling deposit",
     "AS9100 certs", "Free re-sharpening", "Collaborative"),
    ("Brightline Packaging", 2.85, 3.3,
     "Alt converter at 3.10",
     "FSC certified; Net 30; 2-wk lead",
     "Volume tiers; print-plate amortization",
     "FSC certification", "Design support", "Competitive"),
]

OUTCOME_HEADERS = [
    "Supplier", "Date", "Agreed Price", "Target Price", "Walk-Away Price",
    "Savings vs Target", "Savings vs Walk-Away", "Terms Achieved",
    "Concessions Given", "Next Review Date", "Notes",
]

# (supplier, date, agreed, target, walkaway, terms, concessions, next_review, notes)
OUTCOMES = [
    ("Coastal Components Inc.", "2026-07-01", 605, 600, 630,
     "24-mo warranty, Net 60, 98% SLA", "Consigned stock at 2 sites",
     "2027-01-01", "Landed within band; multi-year signed"),
    ("Meridian Metals Co.", "2026-07-03", 210, 205, 222,
     "12-mo price hold, Net 45", "Prompt-pay 1.5%",
     "2026-10-03", "Above target but under walk-away"),
    ("Evergreen Chemicals", "2026-07-05", 135, 128, 145,
     "REACH docs, Net 30, VMI pilot", "Larger drums at same price",
     "2027-07-05", "VMI pilot for 6 months"),
    ("Anchor Logistics LLC", "2026-07-08", 8_500, 8_200, 8_900,
     "97% OTD SLA, fuel cap", "Annual tender commit",
     "2027-07-08", "Fuel-surcharge cap secured"),
    ("Falcon Precision Tooling", "2026-07-10", 46, 44, 50,
     "12-mo warranty, Net 45", "Free re-sharpening x2",
     "2027-01-10", "Blanket PO in place"),
    ("Brightline Packaging", "2026-07-12", 3.05, 2.85, 3.3,
     "FSC, Net 30, volume tiers", "Print-plate amortized over 12 mo",
     "2026-10-12", "Design support included"),
]


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 42)


def build_intel(wb):
    ws = wb.active
    ws.title = "Market Intel"
    ws.append(INTEL_HEADERS)
    for row in MARKET_INTEL:
        ws.append(list(row))
    _style_header(ws, len(INTEL_HEADERS))
    _autosize(ws)


def build_plan(wb):
    ws = wb.create_sheet("Negotiation Plan")
    ws.append(PLAN_HEADERS)
    for row in NEGOTIATION_PLAN:
        ws.append(list(row))
    _style_header(ws, len(PLAN_HEADERS))
    _autosize(ws)


def build_outcome(wb):
    ws = wb.create_sheet("Outcome Log")
    ws.append(OUTCOME_HEADERS)
    for i, o in enumerate(OUTCOMES):
        r = i + 2
        supplier, date, agreed, target, walkaway, terms, concessions, review, notes = o
        ws.cell(row=r, column=1, value=supplier)
        ws.cell(row=r, column=2, value=date)
        ws.cell(row=r, column=3, value=agreed)
        ws.cell(row=r, column=4, value=target)
        ws.cell(row=r, column=5, value=walkaway)
        # Positive savings = agreed came in below target / walk-away.
        ws.cell(row=r, column=6).value = f"=D{r}-C{r}"
        ws.cell(row=r, column=7).value = f"=E{r}-C{r}"
        ws.cell(row=r, column=8, value=terms)
        ws.cell(row=r, column=9, value=concessions)
        ws.cell(row=r, column=10, value=review)
        ws.cell(row=r, column=11, value=notes)
    _style_header(ws, len(OUTCOME_HEADERS))
    _autosize(ws)


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "negotiation_workbook.xlsx"
    wb = Workbook()
    build_intel(wb)
    build_plan(wb)
    build_outcome(wb)
    wb.save(out)
    print(f"Wrote {out} with sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
