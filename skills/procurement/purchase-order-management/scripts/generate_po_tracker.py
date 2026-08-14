#!/usr/bin/env python3
"""Generate a purchase-order tracking workbook.

Builds a 3-sheet .xlsx demonstrating the core data operation behind the
purchase-order-management skill:

  PO Register -> Goods Receipt -> 3-Way Match

- Sheet 1 "PO Register": 12 realistic purchase orders.
- Sheet 2 "Goods Receipt": receiving records against the POs, including
  partial, over, and spoiled receipts.
- Sheet 3 "3-Way Match": PO vs invoice vs receipt with IF-formula match
  status and variance, mixing matched, partial, and mismatched POs.

Run:  python generate_po_tracker.py [output.xlsx]
Requires: openpyxl
"""
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="385723")
HEADER_FONT = Font(bold=True, color="FFFFFF")

PO_HEADERS = [
    "PO ID", "Supplier", "Requisitioner", "Date", "Line Items Count",
    "Total Value", "Status", "Approval Level", "Payment Terms", "Delivery Date",
]

# (po, supplier, requisitioner, date, lines, value, status, approval, terms, delivery)
POS = [
    ("PO-1001", "Meridian Metals Co.", "A. Chen", "2026-06-02", 4, 42_500, "Closed", "Manager", "Net 30", "2026-06-20"),
    ("PO-1002", "Anchor Logistics LLC", "R. Diaz", "2026-06-05", 2, 8_900, "Received", "Lead", "Net 45", "2026-06-18"),
    ("PO-1003", "Brightline Packaging", "M. Osei", "2026-06-07", 6, 15_200, "Closed", "Manager", "Net 30", "2026-06-25"),
    ("PO-1004", "Coastal Components Inc.", "A. Chen", "2026-06-10", 3, 88_000, "Received", "Director", "Net 60", "2026-07-05"),
    ("PO-1005", "Evergreen Chemicals", "T. Novak", "2026-06-12", 1, 5_400, "Sent", "Lead", "Net 30", "2026-06-30"),
    ("PO-1006", "Falcon Precision Tooling", "R. Diaz", "2026-06-14", 5, 23_750, "Received", "Manager", "Net 45", "2026-07-01"),
    ("PO-1007", "Delta Fasteners", "M. Osei", "2026-06-15", 8, 2_150, "Closed", "Lead", "Net 15", "2026-06-24"),
    ("PO-1008", "Meridian Metals Co.", "A. Chen", "2026-06-18", 2, 61_000, "Approved", "Director", "Net 60", "2026-07-10"),
    ("PO-1009", "Anchor Logistics LLC", "T. Novak", "2026-06-20", 1, 3_300, "Draft", "Manager", "Net 45", "2026-07-02"),
    ("PO-1010", "Coastal Components Inc.", "R. Diaz", "2026-06-22", 4, 47_800, "Received", "Manager", "Net 60", "2026-07-08"),
    ("PO-1011", "Brightline Packaging", "M. Osei", "2026-06-24", 3, 9_600, "Sent", "Lead", "Net 30", "2026-07-04"),
    ("PO-1012", "Evergreen Chemicals", "A. Chen", "2026-06-26", 2, 12_400, "Closed", "Manager", "Net 30", "2026-07-12"),
]

GR_HEADERS = [
    "Receipt ID", "PO ID", "Item", "Qty Ordered", "Qty Received",
    "Qty Damaged", "Status", "Receipt Date", "Inspector",
]

# (receipt, po, item, ordered, received, damaged, status, date, inspector)
RECEIPTS = [
    ("GR-5001", "PO-1001", "Steel plate 10mm", 200, 200, 0, "Full", "2026-06-19", "K. Ford"),
    ("GR-5002", "PO-1002", "Freight service", 1, 1, 0, "Full", "2026-06-17", "K. Ford"),
    ("GR-5003", "PO-1003", "Corrugated boxes", 5000, 4800, 0, "Partial", "2026-06-24", "S. Blake"),
    ("GR-5004", "PO-1004", "Control boards", 150, 150, 2, "Full", "2026-07-04", "S. Blake"),
    ("GR-5005", "PO-1006", "CNC end mills", 120, 130, 0, "Over", "2026-06-30", "K. Ford"),
    ("GR-5006", "PO-1007", "Hex bolts M8", 10000, 10000, 0, "Full", "2026-06-23", "J. Ruiz"),
    ("GR-5007", "PO-1010", "Sensor modules", 300, 285, 15, "Spoiled", "2026-07-07", "S. Blake"),
    ("GR-5008", "PO-1012", "Solvent drums", 40, 40, 0, "Full", "2026-07-11", "J. Ruiz"),
]

MATCH_HEADERS = [
    "PO ID", "PO Value", "Invoice Value", "Receipt Value",
    "Match Status", "Variance", "Notes",
]

# (po, po_value, invoice_value, receipt_value, notes)
MATCHES = [
    ("PO-1001", 42_500, 42_500, 42_500, "Clean 3-way match"),
    ("PO-1002", 8_900, 8_900, 8_900, "Service PO, matched"),
    ("PO-1003", 15_200, 15_200, 14_592, "Partial receipt (4800/5000)"),
    ("PO-1004", 88_000, 88_000, 88_000, "Matched; 2 units flagged for QA"),
    ("PO-1006", 23_750, 25_729, 23_750, "Invoice over PO — over-shipment billed"),
    ("PO-1007", 2_150, 2_150, 2_150, "Clean 3-way match"),
    ("PO-1010", 47_800, 47_800, 45_410, "Spoiled units short-received"),
    ("PO-1012", 12_400, 12_400, 12_400, "Clean 3-way match"),
]


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 38)


def build_register(wb):
    ws = wb.active
    ws.title = "PO Register"
    ws.append(PO_HEADERS)
    for row in POS:
        ws.append(list(row))
    for r in range(2, len(POS) + 2):
        ws.cell(row=r, column=6).number_format = "#,##0"
    _style_header(ws, len(PO_HEADERS))
    _autosize(ws)


def build_receipts(wb):
    ws = wb.create_sheet("Goods Receipt")
    ws.append(GR_HEADERS)
    for row in RECEIPTS:
        ws.append(list(row))
    _style_header(ws, len(GR_HEADERS))
    _autosize(ws)


def build_match(wb):
    ws = wb.create_sheet("3-Way Match")
    ws.append(MATCH_HEADERS)
    for i, (po, pov, inv, rec, notes) in enumerate(MATCHES):
        r = i + 2
        ws.cell(row=r, column=1, value=po)
        ws.cell(row=r, column=2, value=pov).number_format = "#,##0"
        ws.cell(row=r, column=3, value=inv).number_format = "#,##0"
        ws.cell(row=r, column=4, value=rec).number_format = "#,##0"
        # Match if PO, invoice, and receipt all agree; else Mismatch.
        ws.cell(row=r, column=5).value = (
            f'=IF(AND(B{r}=C{r},B{r}=D{r}),"Match","Mismatch")'
        )
        # Variance = invoice value - receipt value (what may be over/under billed).
        ws.cell(row=r, column=6).value = f"=C{r}-D{r}"
        ws.cell(row=r, column=6).number_format = "#,##0"
        ws.cell(row=r, column=7, value=notes)
    _style_header(ws, len(MATCH_HEADERS))
    _autosize(ws)

    n = len(MATCHES)
    status_col = get_column_letter(5)
    rng = f"{status_col}2:{status_col}{n + 1}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Mismatch"'],
                        fill=PatternFill("solid", fgColor="FFC7CE")))
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Match"'],
                        fill=PatternFill("solid", fgColor="C6EFCE")))


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "po_tracker.xlsx"
    wb = Workbook()
    build_register(wb)
    build_receipts(wb)
    build_match(wb)
    wb.save(out)
    print(f"Wrote {out} with sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
