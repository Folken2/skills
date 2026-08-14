#!/usr/bin/env python3
"""Generate a supplier qualification workbook.

Builds a 3-sheet .xlsx demonstrating the core data operation behind the
supplier-qualification skill:

  Supplier Intake -> weighted Scorecard -> Approved List

- Sheet 1 "Supplier Intake": 8 realistic supplier candidates.
- Sheet 2 "Scorecard": weighted-criteria scores with a formula-driven total
  and a tier (A/B/C/D), colour-coded by conditional formatting.
- Sheet 3 "Approved List": suppliers with Total Score >= 70, sorted desc.

Run:  python generate_supplier_scorecard.py [output.xlsx]
Requires: openpyxl
"""
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# Weighted criteria: (label, weight). Weights sum to 1.00.
CRITERIA = [
    ("Financial Health", 0.25),
    ("Quality", 0.25),
    ("Capacity", 0.15),
    ("Compliance", 0.15),
    ("Risk", 0.10),
    ("References", 0.10),
]

# Supplier intake rows and per-criterion raw scores (0-100) used for the
# weighted scorecard. Column order of the scores matches CRITERIA.
SUPPLIERS = [
    # (id, name, category, location, contact, revenue, years, certs, insurance, refs), [scores]
    ("SUP-001", "Meridian Metals Co.", "Raw Materials", "Pittsburgh, PA", "j.hale@meridian.com",
     48_000_000, 22, "ISO 9001, ISO 14001", "$5M GL", 3, [88, 92, 80, 90, 85, 90]),
    ("SUP-002", "Anchor Logistics LLC", "Logistics", "Memphis, TN", "ops@anchorlog.com",
     120_000_000, 15, "ISO 9001, C-TPAT", "$10M GL", 4, [82, 78, 88, 84, 70, 80]),
    ("SUP-003", "Brightline Packaging", "Packaging", "Columbus, OH", "sales@brightline.com",
     9_500_000, 8, "ISO 9001, FSC", "$2M GL", 2, [70, 74, 65, 72, 68, 70]),
    ("SUP-004", "Coastal Components Inc.", "Electronics", "San Jose, CA", "rfq@coastalcomp.com",
     205_000_000, 30, "ISO 9001, IATF 16949, RoHS", "$15M GL", 5, [95, 93, 90, 94, 90, 95]),
    ("SUP-005", "Delta Fasteners", "Hardware", "Cleveland, OH", "orders@deltafast.com",
     3_200_000, 5, "ISO 9001", "$1M GL", 1, [55, 60, 50, 58, 52, 48]),
    ("SUP-006", "Evergreen Chemicals", "Chemicals", "Houston, TX", "supply@evergreenchem.com",
     78_000_000, 18, "ISO 9001, ISO 45001, REACH", "$8M GL", 3, [80, 85, 78, 88, 60, 82]),
    ("SUP-007", "Falcon Precision Tooling", "Machining", "Grand Rapids, MI", "quote@falcontool.com",
     14_000_000, 11, "ISO 9001, AS9100", "$3M GL", 3, [76, 88, 70, 80, 74, 72]),
    ("SUP-008", "GreyRock Textiles", "Textiles", "Charlotte, NC", "hello@greyrock.com",
     1_800_000, 3, "None", "$0.5M GL", 0, [40, 45, 42, 38, 35, 30]),
]

INTAKE_HEADERS = [
    "Supplier ID", "Name", "Category", "Location", "Contact",
    "Annual Revenue", "Years in Business", "Certifications",
    "Insurance Coverage", "References",
]


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 40)


def build_intake(wb):
    ws = wb.active
    ws.title = "Supplier Intake"
    ws.append(INTAKE_HEADERS)
    for row in SUPPLIERS:
        ws.append(list(row[:10]))
    for r in range(2, len(SUPPLIERS) + 2):
        ws.cell(row=r, column=6).number_format = "#,##0"
    _style_header(ws, len(INTAKE_HEADERS))
    _autosize(ws)
    return ws


def build_scorecard(wb):
    ws = wb.create_sheet("Scorecard")
    labels = [f"{lbl} ({int(w * 100)}%)" for lbl, w in CRITERIA]
    headers = ["Supplier ID", "Name"] + labels + ["Total Score", "Tier"]
    ws.append(headers)

    n = len(SUPPLIERS)
    first_score_col = 3  # column C
    last_score_col = first_score_col + len(CRITERIA) - 1
    total_col = last_score_col + 1
    tier_col = total_col + 1

    for i, row in enumerate(SUPPLIERS):
        sup, scores = row[:10], row[10]
        r = i + 2
        ws.cell(row=r, column=1, value=sup[0])
        ws.cell(row=r, column=2, value=sup[1])
        for j, val in enumerate(scores):
            ws.cell(row=r, column=first_score_col + j, value=val)
        # Weighted total via SUMPRODUCT of the score range against the weights.
        score_range = (
            f"{get_column_letter(first_score_col)}{r}:"
            f"{get_column_letter(last_score_col)}{r}"
        )
        weights = ",".join(str(w) for _, w in CRITERIA)
        total_cell = ws.cell(row=r, column=total_col)
        total_cell.value = f"=SUMPRODUCT({score_range},{{{weights}}})"
        total_cell.number_format = "0.0"
        tl = get_column_letter(total_col)
        ws.cell(row=r, column=tier_col).value = (
            f'=IF({tl}{r}>85,"A",IF({tl}{r}>=70,"B",IF({tl}{r}>=50,"C","D")))'
        )

    _style_header(ws, len(headers))
    _autosize(ws)

    # Conditional formatting on the Total Score column by tier band.
    total_letter = get_column_letter(total_col)
    rng = f"{total_letter}2:{total_letter}{n + 1}"
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    orange = PatternFill("solid", fgColor="FCD5B4")
    red = PatternFill("solid", fgColor="FFC7CE")
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["85"], fill=green))
    ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["70", "85"], fill=yellow))
    ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["50", "69.9999"], fill=orange))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["50"], fill=red))
    return ws


def build_approved_list(wb):
    ws = wb.create_sheet("Approved List")
    ws.append(["Supplier ID", "Name", "Category", "Total Score", "Tier"])

    # Compute totals in Python so we can filter/sort the static approved list.
    def total(scores):
        return round(sum(s * w for s, (_, w) in zip(scores, CRITERIA)), 1)

    def tier(t):
        return "A" if t > 85 else "B" if t >= 70 else "C" if t >= 50 else "D"

    scored = [
        (row[0], row[1], row[2], total(row[10]))
        for row in SUPPLIERS
    ]
    approved = sorted(
        [row for row in scored if row[3] >= 70], key=lambda x: x[3], reverse=True
    )
    for sid, name, cat, tot in approved:
        ws.append([sid, name, cat, tot, tier(tot)])
    _style_header(ws, 5)
    _autosize(ws)
    return ws


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "supplier_scorecard.xlsx"
    wb = Workbook()
    build_intake(wb)
    build_scorecard(wb)
    build_approved_list(wb)
    wb.save(out)
    print(f"Wrote {out} with sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
