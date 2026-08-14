#!/usr/bin/env python3
"""Generate an inventory-planning workbook.

Builds a 3-sheet .xlsx demonstrating the core data operation behind the
inventory-planning skill:

  Inventory Snapshot -> Replenishment Plan -> ABC Analysis

- Sheet 1 "Inventory Snapshot": 15 SKUs across 3 categories.
- Sheet 2 "Replenishment Plan": reorder point, safety stock, EOQ, and
  suggested order — ALL computed with formulas that reference the snapshot —
  with conditional formatting on the stock-vs-reorder flag.
- Sheet 3 "ABC Analysis": ranks SKUs by annual consumption value and
  classifies A (top 80%), B (next 15%), C (bottom 5%) via formulas.

Formulas used:
  Safety Stock = Lead Time Variability * Monthly Demand / 30 * 1.65
  Reorder Point = Monthly Demand / 30 * Lead Time + Safety Stock
  EOQ = SQRT(2 * AnnualDemand * OrderCost / (HoldingRate * Unit Cost))

Run:  python generate_inventory_plan.py [output.xlsx]
Requires: openpyxl
"""
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="203864")
HEADER_FONT = Font(bold=True, color="FFFFFF")

ORDER_COST = 50.0      # fixed cost per order, used in EOQ
HOLDING_RATE = 0.25    # annual holding cost as a fraction of unit cost

SNAP_HEADERS = [
    "SKU", "Product", "Category", "Current Stock", "Unit Cost",
    "Monthly Demand", "Lead Time (days)", "Lead Time Variability (days)", "Supplier",
]

# (sku, product, category, current_stock, unit_cost, monthly_demand, lead_time, lt_var, supplier)
SKUS = [
    ("SKU-A100", "Control board rev C", "Electronics", 120, 590.0, 90, 21, 5, "Coastal Components Inc."),
    ("SKU-A101", "Sensor module", "Electronics", 40, 145.0, 300, 18, 4, "Coastal Components Inc."),
    ("SKU-A102", "Wiring harness", "Electronics", 500, 22.5, 800, 14, 3, "Falcon Precision Tooling"),
    ("SKU-A103", "LCD panel 7in", "Electronics", 25, 38.0, 150, 30, 8, "Coastal Components Inc."),
    ("SKU-A104", "Power supply 120W", "Electronics", 60, 64.0, 120, 21, 6, "Coastal Components Inc."),
    ("SKU-B200", "Steel plate 10mm", "Raw Materials", 200, 210.0, 60, 35, 10, "Meridian Metals Co."),
    ("SKU-B201", "Aluminum billet", "Raw Materials", 80, 175.0, 45, 28, 7, "Meridian Metals Co."),
    ("SKU-B202", "Copper wire spool", "Raw Materials", 340, 48.0, 220, 21, 5, "Meridian Metals Co."),
    ("SKU-B203", "Hex bolts M8", "Raw Materials", 12000, 0.12, 10000, 15, 3, "Delta Fasteners"),
    ("SKU-B204", "Industrial solvent", "Raw Materials", 90, 132.0, 40, 14, 4, "Evergreen Chemicals"),
    ("SKU-C300", "Corrugated boxes", "Packaging", 4800, 2.9, 5000, 21, 6, "Brightline Packaging"),
    ("SKU-C301", "Stretch wrap roll", "Packaging", 220, 14.0, 400, 10, 2, "Brightline Packaging"),
    ("SKU-C302", "Pallet 48x40", "Packaging", 150, 18.5, 300, 14, 4, "Brightline Packaging"),
    ("SKU-C303", "Void-fill paper", "Packaging", 60, 26.0, 180, 12, 3, "Brightline Packaging"),
    ("SKU-C304", "Shipping labels", "Packaging", 30000, 0.02, 40000, 7, 2, "Brightline Packaging"),
]


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 34)


def build_snapshot(wb):
    ws = wb.active
    ws.title = "Inventory Snapshot"
    ws.append(SNAP_HEADERS)
    for row in SKUS:
        ws.append(list(row))
    for r in range(2, len(SKUS) + 2):
        ws.cell(row=r, column=5).number_format = "#,##0.00"
    _style_header(ws, len(SNAP_HEADERS))
    _autosize(ws)


def build_replenishment(wb):
    ws = wb.create_sheet("Replenishment Plan")
    headers = [
        "SKU", "Current Stock", "Safety Stock", "Reorder Point", "EOQ",
        "Suggested Order", "Stock vs Reorder",
    ]
    ws.append(headers)
    snap = "'Inventory Snapshot'"
    n = len(SKUS)
    for i in range(n):
        r = i + 2  # same row alignment as the snapshot sheet
        # Snapshot refs: D=current, E=unit cost, F=monthly demand,
        #                G=lead time, H=lead-time variability.
        cur = f"{snap}!D{r}"
        cost = f"{snap}!E{r}"
        dem = f"{snap}!F{r}"
        lt = f"{snap}!G{r}"
        ltv = f"{snap}!H{r}"

        ws.cell(row=r, column=1).value = f"{snap}!A{r}"
        ws.cell(row=r, column=2).value = f"={cur}"
        # Safety Stock = LT variability * monthly demand / 30 * 1.65
        ws.cell(row=r, column=3).value = f"=ROUND({ltv}*{dem}/30*1.65,0)"
        # Reorder Point = monthly demand / 30 * lead time + safety stock
        ws.cell(row=r, column=4).value = f"=ROUND({dem}/30*{lt}+C{r},0)"
        # EOQ = sqrt(2 * annual demand * order cost / (holding rate * unit cost))
        ws.cell(row=r, column=5).value = (
            f"=ROUND(SQRT(2*({dem}*12)*{ORDER_COST}/({HOLDING_RATE}*{cost})),0)"
        )
        # Suggested order: reorder up to ROP+EOQ when below reorder point.
        ws.cell(row=r, column=6).value = f"=IF(B{r}<D{r},D{r}+E{r}-B{r},0)"
        # Flag current vs reorder point.
        ws.cell(row=r, column=7).value = (
            f'=IF(B{r}<D{r},"BELOW",IF(B{r}<=D{r}*1.2,"NEAR","OK"))'
        )
    _style_header(ws, len(headers))
    _autosize(ws)

    rng = f"G2:G{n + 1}"
    # Red when below reorder point, yellow when within 20% of it.
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=['$G2="BELOW"'],
                         fill=PatternFill("solid", fgColor="FFC7CE")))
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=['$G2="NEAR"'],
                         fill=PatternFill("solid", fgColor="FFEB9C")))
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=['$G2="OK"'],
                         fill=PatternFill("solid", fgColor="C6EFCE")))


def build_abc(wb):
    ws = wb.create_sheet("ABC Analysis")
    headers = [
        "SKU", "Product", "Annual Consumption Value", "Cumulative Value",
        "Cumulative %", "Class",
    ]
    ws.append(headers)
    snap = "'Inventory Snapshot'"

    # Rank SKUs by annual consumption value (monthly demand * 12 * unit cost)
    # descending so the cumulative-% formula produces a clean Pareto ranking.
    ranked = sorted(
        enumerate(SKUS),
        key=lambda kv: kv[1][5] * 12 * kv[1][4],
        reverse=True,
    )
    n = len(ranked)
    for out_i, (orig_i, sku) in enumerate(ranked):
        r = out_i + 2
        src = orig_i + 2  # snapshot row for this SKU
        ws.cell(row=r, column=1).value = f"={snap}!A{src}"
        ws.cell(row=r, column=2).value = f"={snap}!B{src}"
        # Annual consumption value = monthly demand * 12 * unit cost.
        ws.cell(row=r, column=3).value = f"={snap}!F{src}*12*{snap}!E{src}"
        ws.cell(row=r, column=3).number_format = "#,##0.00"
        # Running cumulative value from the top of the ranked list.
        ws.cell(row=r, column=4).value = f"=SUM($C$2:C{r})"
        ws.cell(row=r, column=4).number_format = "#,##0.00"
        # Cumulative % of total annual value.
        ws.cell(row=r, column=5).value = f"=D{r}/SUM($C$2:$C${n + 1})"
        ws.cell(row=r, column=5).number_format = "0.0%"
        # Class A up to 80% cumulative, B up to 95%, C thereafter.
        ws.cell(row=r, column=6).value = (
            f'=IF(E{r}<=0.8,"A",IF(E{r}<=0.95,"B","C"))'
        )
    _style_header(ws, len(headers))
    _autosize(ws)


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "inventory_plan.xlsx"
    wb = Workbook()
    build_snapshot(wb)
    build_replenishment(wb)
    build_abc(wb)
    wb.save(out)
    print(f"Wrote {out} with sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
