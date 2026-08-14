#!/usr/bin/env python3
"""Generate an order-management workbook.

Builds a three-sheet .xlsx demonstrating the core order data operations:
looking up an order, checking status, and tracking returns/refunds.

    Sheet 1 "Orders Master"   - 10 orders in various statuses
    Sheet 2 "Returns/Refunds" - return/refund tracking rows
    Sheet 3 "Order Lookup"    - enter an Order ID, XLOOKUP/VLOOKUP pulls details

Run:  python3 generate_order_template.py [output.xlsx]
Requires: openpyxl  (pip install openpyxl)
"""
import sys
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F6F43")

ORDER_HEADERS = [
    "Order ID", "Customer", "Date", "Product", "Qty", "Total",
    "Status", "Tracking", "Est Delivery", "Actual Delivery",
]

# (customer, product, qty, total, status, tracking, est_offset_days, actual_offset_days)
# actual_offset_days = None means not yet delivered.
ORDERS = [
    ("Jane Fielding",  "Wireless Headphones",     1,  129.99, "Delivered",  "1Z9990001", 3,  3),
    ("Bright Media",   "4K Monitor 27\"",         2,  639.98, "Shipped",    "1Z9990002", 5,  None),
    ("Nova Labs",      "USB-C Dock",              1,   89.00, "Processing", "",          6,  None),
    ("Tom Reyes",      "Mechanical Keyboard",     1,  119.50, "Delivered",  "1Z9990004", 2,  4),
    ("Priya Nair",     "Ergonomic Chair",         1,  349.00, "Returned",   "1Z9990005", 4,  4),
    ("Delta Freight",  "Label Printer",           3,  447.00, "Disputed",   "1Z9990006", 5,  6),
    ("Kenji Watanabe", "Phone Case (Blue)",       2,   39.98, "Delivered",  "1Z9990007", 3,  3),
    ("Sunrise Retail", "Barcode Scanner",         5,  795.00, "Shipped",    "1Z9990008", 7,  None),
    ("Grace Oduya",    "Standing Desk",           1,  429.00, "Processing", "",          8,  None),
    ("Acme Corp",      "Webcam 1080p",            4,  239.96, "Returned",   "1Z9990010", 3,  2),
]

RETURN_HEADERS = [
    "Return ID", "Order ID", "Reason", "Amount", "Refund Status",
    "Date Initiated", "Date Completed", "Notes",
]


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def build_orders_sheet(wb):
    ws = wb.active
    ws.title = "Orders Master"
    ws.append(ORDER_HEADERS)
    _style_header(ws, len(ORDER_HEADERS))

    base = datetime(2026, 8, 1)
    for i, (cust, product, qty, total, status, tracking, est, actual) in enumerate(ORDERS, start=1):
        order_date = base + timedelta(days=i)
        est_delivery = order_date + timedelta(days=est)
        actual_delivery = order_date + timedelta(days=actual) if actual is not None else ""
        ws.append([
            f"ORD-{2000 + i}",
            cust,
            order_date.strftime("%Y-%m-%d"),
            product,
            qty,
            total,
            status,
            tracking,
            est_delivery.strftime("%Y-%m-%d"),
            actual_delivery.strftime("%Y-%m-%d") if actual_delivery else "",
        ])
    # currency format on Total column (F)
    for row in range(2, len(ORDERS) + 2):
        ws.cell(row=row, column=6).number_format = '"$"#,##0.00'
    for idx, w in enumerate([10, 16, 12, 22, 6, 12, 12, 12, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    return ws


def build_returns_sheet(wb):
    # Excel forbids "/" in sheet titles, so the Returns/Refunds tracker uses a hyphen.
    ws = wb.create_sheet("Returns-Refunds")
    ws.append(RETURN_HEADERS)
    _style_header(ws, len(RETURN_HEADERS))
    returns = [
        ("RET-01", "ORD-2005", "Uncomfortable / buyer's remorse", 349.00, "Completed",
         "2026-08-10", "2026-08-13", "Standard 30-day return, card refund"),
        ("RET-02", "ORD-2010", "Wrong item shipped (company fault)", 239.96, "Completed",
         "2026-08-06", "2026-08-07", "Reship + full refund, no cost to customer"),
        ("RET-03", "ORD-2006", "Damaged on arrival - disputed", 447.00, "Pending",
         "2026-08-12", "", "Awaiting carrier trace before refund"),
    ]
    for r in returns:
        ws.append(list(r))
    for row in range(2, len(returns) + 2):
        ws.cell(row=row, column=4).number_format = '"$"#,##0.00'
    for idx, w in enumerate([10, 10, 34, 12, 14, 14, 14, 40], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    return ws


def build_lookup_sheet(wb, n_orders):
    ws = wb.create_sheet("Order Lookup")
    ws["A1"] = "Order Lookup"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Enter Order ID:"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = "ORD-2005"  # sample query
    ws["B3"].fill = PatternFill("solid", fgColor="FFF2CC")

    last = n_orders + 1
    # Prefer XLOOKUP; fall back to VLOOKUP for older engines (both stored; XLOOKUP wins where supported).
    fields = [
        ("Customer", 2),
        ("Date", 3),
        ("Product", 4),
        ("Qty", 5),
        ("Total", 6),
        ("Status", 7),
        ("Tracking", 8),
        ("Est Delivery", 9),
        ("Actual Delivery", 10),
    ]
    for offset, (label, col) in enumerate(fields, start=5):
        ws.cell(row=offset, column=1, value=label).font = Font(bold=True)
        col_letter = get_column_letter(col)
        formula = (
            f"=IFERROR(XLOOKUP($B$3,'Orders Master'!$A$2:$A${last},"
            f"'Orders Master'!${col_letter}$2:${col_letter}${last}),"
            f"VLOOKUP($B$3,'Orders Master'!$A$2:$J${last},{col},FALSE))"
        )
        ws.cell(row=offset, column=2, value=formula)
    ws.cell(row=9, column=2).number_format = '"$"#,##0.00'  # Total row
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 24
    return ws


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "order_management_template.xlsx"
    wb = Workbook()
    build_orders_sheet(wb)
    build_returns_sheet(wb)
    build_lookup_sheet(wb, len(ORDERS))
    wb.save(out)
    print(f"Wrote {out} with sheets: Orders Master, Returns/Refunds, Order Lookup "
          f"({len(ORDERS)} sample orders)")


if __name__ == "__main__":
    main()
