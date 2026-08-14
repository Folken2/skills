#!/usr/bin/env python3
"""Generate a ticket-triage tracking workbook.

Builds a three-sheet .xlsx demonstrating the core triage data operation:
reading raw tickets, classifying them, and reporting on the queue.

    Sheet 1 "Tickets"     - one row per ticket with severity/tier/SLA/priority
    Sheet 2 "SLA Matrix"  - Issue Type x Severity -> response/resolution SLA hours
    Sheet 3 "Daily Report" - COUNTIFS summary of open/closed/escalated

Run:  python3 generate_ticket_template.py [output.xlsx]
Requires: openpyxl  (pip install openpyxl)
"""
import sys
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Priority score by severity, and tier routing by issue type.
PRIORITY_BY_SEVERITY = {"Critical": 100, "High": 75, "Medium": 50, "Low": 25}
TIER_BY_TYPE = {
    "billing": "Tier1",
    "question": "Tier1",
    "technical": "Tier2",
    "order": "Tier2",
    "bug": "Tier3",
}
# Response / resolution SLA (hours) keyed by (issue_type_class, severity).
# issue_type_class collapses the raw type into billing/technical/bug for the matrix.
SLA_MATRIX = {
    ("billing", "Critical"): (1, 8),
    ("billing", "High"): (2, 24),
    ("billing", "Medium"): (8, 72),
    ("billing", "Low"): (24, 120),
    ("technical", "Critical"): (1, 4),
    ("technical", "High"): (2, 12),
    ("technical", "Medium"): (8, 48),
    ("technical", "Low"): (24, 96),
    ("bug", "Critical"): (1, 8),
    ("bug", "High"): (4, 24),
    ("bug", "Medium"): (12, 72),
    ("bug", "Low"): (48, 168),
}

# 10 realistic raw tickets: (customer, issue_text, issue_type, severity, status)
RAW_TICKETS = [
    ("Acme Corp",        "Login page returns 500 error for all SSO users",        "technical", "Critical", "Escalated"),
    ("Jane Fielding",    "Order #1234 still not delivered after 2 weeks",          "order",     "High",     "Open"),
    ("Bright Media",     "Double charged on invoice INV-8821",                    "billing",   "High",     "Open"),
    ("Tom Reyes",        "How do I export my report to CSV?",                     "question",  "Low",      "Closed"),
    ("Nova Labs",        "Dashboard graphs render blank on Safari",               "bug",       "Medium",   "Open"),
    ("Priya Nair",       "Refund for cancelled subscription not received",        "billing",   "Medium",   "Open"),
    ("Delta Freight",    "API returns 429 rate-limit under normal load",          "technical", "High",     "Escalated"),
    ("Kenji Watanabe",   "Mobile app crashes on photo upload",                   "bug",       "High",     "Open"),
    ("Sunrise Retail",   "Wrong item shipped for order #5567",                    "order",     "Medium",   "Open"),
    ("Grace Oduya",      "Can I change the email on my account?",                 "question",  "Low",      "Closed"),
]

TICKET_HEADERS = [
    "Ticket ID", "Date", "Customer", "Issue Type", "Severity",
    "Status", "Priority Score", "Assigned Tier", "SLA Due", "Notes",
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")


def _classify(issue_type):
    """Collapse a raw issue type into the SLA matrix class."""
    if issue_type in ("billing",):
        return "billing"
    if issue_type in ("bug",):
        return "bug"
    return "technical"  # technical, order, question all route through the technical SLA


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def build_tickets_sheet(wb):
    ws = wb.active
    ws.title = "Tickets"
    ws.append(TICKET_HEADERS)
    _style_header(ws, len(TICKET_HEADERS))

    base = datetime(2026, 8, 14, 9, 0)
    for i, (customer, issue, itype, severity, status) in enumerate(RAW_TICKETS, start=1):
        created = base - timedelta(hours=i * 3)
        cls = _classify(itype)
        response_sla, _resolution_sla = SLA_MATRIX[(cls, severity)]
        sla_due = created + timedelta(hours=response_sla)
        priority = PRIORITY_BY_SEVERITY[severity]
        tier = TIER_BY_TYPE[itype]
        ws.append([
            f"TKT-{1000 + i}",
            created.strftime("%Y-%m-%d %H:%M"),
            customer,
            itype,
            severity,
            status,
            priority,
            tier,
            sla_due.strftime("%Y-%m-%d %H:%M"),
            issue,
        ])

    widths = [10, 16, 16, 12, 10, 10, 14, 12, 16, 48]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    return ws


def build_sla_matrix_sheet(wb):
    ws = wb.create_sheet("SLA Matrix")
    ws.append(["Issue Type Class", "Severity", "Response SLA (hrs)", "Resolution SLA (hrs)"])
    _style_header(ws, 4)
    for (cls, severity), (resp, res) in SLA_MATRIX.items():
        ws.append([cls, severity, resp, res])
    for idx, w in enumerate([18, 12, 20, 22], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    return ws


def build_daily_report_sheet(wb, n_tickets):
    ws = wb.create_sheet("Daily Report")
    ws["A1"] = "Daily Triage Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Metric"
    ws["B3"] = "Count"
    for c in ("A3", "B3"):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL

    last = n_tickets + 1  # data rows are 2..last on the Tickets sheet
    rows = [
        ("Open tickets",       f'=COUNTIF(Tickets!F2:F{last},"Open")'),
        ("Closed tickets",     f'=COUNTIF(Tickets!F2:F{last},"Closed")'),
        ("Escalated tickets",  f'=COUNTIF(Tickets!F2:F{last},"Escalated")'),
        ("Critical severity",  f'=COUNTIF(Tickets!E2:E{last},"Critical")'),
        ("Tier3 assigned",     f'=COUNTIF(Tickets!H2:H{last},"Tier3")'),
        ("Total tickets",      f"=COUNTA(Tickets!A2:A{last})"),
    ]
    for offset, (label, formula) in enumerate(rows, start=4):
        ws.cell(row=offset, column=1, value=label)
        ws.cell(row=offset, column=2, value=formula)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    return ws


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "ticket_triage_template.xlsx"
    wb = Workbook()
    build_tickets_sheet(wb)
    build_sla_matrix_sheet(wb)
    build_daily_report_sheet(wb, len(RAW_TICKETS))
    wb.save(out)
    print(f"Wrote {out} with sheets: Tickets, SLA Matrix, Daily Report "
          f"({len(RAW_TICKETS)} sample tickets)")


if __name__ == "__main__":
    main()
