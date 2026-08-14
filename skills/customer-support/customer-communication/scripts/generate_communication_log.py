#!/usr/bin/env python3
"""Generate a customer-communication workbook.

Builds a two-sheet .xlsx demonstrating the core communication data ops:
reviewing contact history and pulling the right response template.

    Sheet 1 "Contact Log"       - 10 logged customer contacts
    Sheet 2 "Response Templates" - reusable templates by scenario/channel/tone

Run:  python3 generate_communication_log.py [output.xlsx]
Requires: openpyxl  (pip install openpyxl)
"""
import sys
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="C0504D")

LOG_HEADERS = [
    "Contact ID", "Customer", "Channel", "Issue Type", "Tone Used",
    "Outcome", "Follow-up Required", "Notes",
]

# (customer, channel, issue_type, tone, outcome, followup, notes)
CONTACTS = [
    ("Jane Fielding",  "Email", "Delivery delay",  "Empathetic",  "Reassured, new ETA given", "Yes",
     "Promised update by Fri"),
    ("Bright Media",   "Chat",  "Billing dispute", "Professional", "Refund initiated",         "Yes",
     "Awaiting card settlement"),
    ("Nova Labs",      "Email", "Bug report",      "Professional", "Escalated to eng",         "Yes",
     "Ticket TKT-1004 linked"),
    ("Tom Reyes",      "Chat",  "How-to question", "Professional", "Resolved on first contact", "No",
     "Sent CSV export steps"),
    ("Priya Nair",     "Phone", "Refund status",   "Empathetic",  "Explained timeline",       "No",
     "Refund on day 5"),
    ("Delta Freight",  "Email", "Complaint",       "Empathetic",  "Apology + goodwill credit", "Yes",
     "Manager CC'd"),
    ("Kenji Watanabe", "Chat",  "App crash",       "Urgent",      "Workaround provided",      "Yes",
     "Fix ETA next release"),
    ("Sunrise Retail", "Phone", "Wrong item",      "Empathetic",  "Reship arranged",          "No",
     "Prepaid return label sent"),
    ("Grace Oduya",    "Email", "Account change",  "Professional", "Verified + updated",       "No",
     "Email change confirmed"),
    ("Acme Corp",      "Email", "Positive feedback", "Professional", "Thanked, logged NPS",   "No",
     "Referral offer sent"),
]

TEMPLATE_HEADERS = [
    "Scenario Type", "Channel", "Tone", "Template Text", "Key Phrases",
]
TEMPLATES = [
    ("Complaint", "Email", "Empathetic",
     "Hi {name}, I'm sorry this fell short — that's not the experience we want for you. "
     "Here's what I'm doing to make it right: {remedy}. You'll see {next_step} by {date}.",
     "I'm sorry; make it right; here's what I'm doing"),
    ("Delivery delay", "Email", "Empathetic",
     "Hi {name}, your order {order_id} is running behind. The updated delivery estimate is {new_eta}. "
     "I'll keep you posted and you can reach me directly if anything changes.",
     "updated estimate; I'll keep you posted"),
    ("Refund", "Chat", "Professional",
     "Thanks {name} — I've initiated your refund of {amount}. Card refunds settle in {window} business days. "
     "I'll confirm here once it completes.",
     "I've initiated; settles in; I'll confirm"),
    ("Denial / saying no", "Email", "Professional",
     "Hi {name}, I understand why you'd want {request}. We're not able to do that because {reason}. "
     "What I can offer instead is {alternative}.",
     "I understand; not able to; what I can offer"),
    ("Escalation", "Email", "Urgent",
     "Hi {name}, I'm bringing in my manager so we get this resolved properly. "
     "{owner} will follow up by {date} with the next steps.",
     "bringing in my manager; follow up by"),
    ("Positive follow-up", "Email", "Professional",
     "Hi {name}, just checking that {resolution} is still holding up well. "
     "Thanks again for your patience — reach out anytime.",
     "just checking; thanks again; reach out anytime"),
    ("Pricing objection", "Chat", "Professional",
     "I hear you on the cost, {name}. Here's the value it covers: {value}. "
     "If budget is the constraint, {alternative} may fit better.",
     "I hear you; value it covers; may fit better"),
]


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def build_contact_log(wb):
    ws = wb.active
    ws.title = "Contact Log"
    ws.append(LOG_HEADERS)
    _style_header(ws, len(LOG_HEADERS))
    base = datetime(2026, 8, 13, 9, 0)
    for i, (cust, channel, issue, tone, outcome, followup, notes) in enumerate(CONTACTS, start=1):
        when = base + timedelta(hours=i)
        ws.append([
            f"CON-{500 + i}", cust, channel, issue, tone, outcome, followup,
            f"[{when.strftime('%Y-%m-%d %H:%M')}] {notes}",
        ])
    for idx, w in enumerate([11, 16, 10, 18, 14, 26, 16, 34], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    return ws


def build_templates(wb):
    ws = wb.create_sheet("Response Templates")
    ws.append(TEMPLATE_HEADERS)
    _style_header(ws, len(TEMPLATE_HEADERS))
    for row in TEMPLATES:
        ws.append(list(row))
    wrap = Alignment(wrap_text=True, vertical="top")
    for r in range(2, len(TEMPLATES) + 2):
        ws.cell(row=r, column=4).alignment = wrap
        ws.cell(row=r, column=5).alignment = wrap
    for idx, w in enumerate([18, 10, 14, 60, 34], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    return ws


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "communication_log.xlsx"
    wb = Workbook()
    build_contact_log(wb)
    build_templates(wb)
    wb.save(out)
    print(f"Wrote {out} with sheets: Contact Log, Response Templates "
          f"({len(CONTACTS)} contacts, {len(TEMPLATES)} templates)")


if __name__ == "__main__":
    main()
