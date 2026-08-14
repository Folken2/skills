#!/usr/bin/env python3
"""Generate a technical-troubleshooting workbook.

Builds a three-sheet .xlsx demonstrating the core troubleshooting data ops:
logging cases, searching a KB, and reporting weekly trends.

    Sheet 1 "Troubleshooting Log" - one row per case (symptom -> resolution)
    Sheet 2 "KB Resolutions"      - known symptom patterns and fixes
    Sheet 3 "Weekly Trend"        - COUNTIF trends and escalation rate

Run:  python3 generate_troubleshooting_log.py [output.xlsx]
Requires: openpyxl  (pip install openpyxl)
"""
import sys
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="7030A0")

LOG_HEADERS = [
    "Case ID", "Date", "Product", "Issue", "Symptom", "Diagnosis",
    "Resolution", "Status", "Time Spent (mins)", "Escalated (Y/N)",
]

# (product, issue, symptom, diagnosis, resolution, status, minutes, escalated)
CASES = [
    ("API Gateway", "Cannot connect to API", "Connection refused on :443",
     "Expired TLS cert on edge node", "Rotated cert, redeployed edge", "Resolved", 45, "N"),
    ("Mobile App", "Feature X not working on mobile", "Button greyed out on iOS 17",
     "Feature flag off for iOS build", "Enabled flag, forced config refresh", "Resolved", 30, "N"),
    ("Web App", "Dashboard slow to load", "8s load on reports page",
     "N+1 query on widgets", "Escalated to eng for query fix", "Escalated", 60, "Y"),
    ("Billing", "Invoice PDF blank", "Downloaded PDF has no line items",
     "Template render race condition", "Applied retry, verified output", "Resolved", 40, "N"),
    ("API Gateway", "429 under normal load", "Rate limit at 50 req/s",
     "Misconfigured tenant quota", "Raised quota to plan limit", "Resolved", 25, "N"),
    ("Mobile App", "App crashes on photo upload", "Crash on >10MB images",
     "Unbounded memory on decode", "Escalated to eng, added workaround", "Escalated", 75, "Y"),
    ("Web App", "SSO login loop", "Redirects back to login",
     "Clock skew on SAML assertion", "Synced NTP on IdP node", "Resolved", 50, "N"),
    ("Integrations", "Webhook not firing", "No events received",
     "Endpoint returned 500, disabled", "Fixed endpoint, re-enabled hook", "Resolved", 35, "N"),
]

KB_HEADERS = ["Issue Type", "Symptom Pattern", "Resolution Steps", "References"]
KB = [
    ("Connectivity", "Connection refused / timeout on API",
     "1) Check cert expiry 2) Verify DNS 3) Confirm firewall rules", "KB-101, RUNBOOK-net"),
    ("Auth", "SSO redirect loop / login bounce",
     "1) Check clock skew 2) Validate SAML metadata 3) Clear session", "KB-114"),
    ("Performance", "Page load > 5s",
     "1) Check slow query log 2) Profile N+1 3) Verify cache hit rate", "KB-120, DASH-perf"),
    ("Rate Limiting", "HTTP 429 under normal use",
     "1) Check tenant quota 2) Compare to plan 3) Raise or throttle", "KB-131"),
    ("Rendering", "PDF/export blank or partial",
     "1) Retry render 2) Check template job logs 3) Verify data source", "KB-140"),
    ("Mobile", "Crash on media upload",
     "1) Check file size cap 2) Reproduce on device 3) Escalate if OOM", "KB-155"),
    ("Integrations", "Webhook events missing",
     "1) Check endpoint health 2) Inspect delivery log 3) Re-enable hook", "KB-162"),
]


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def build_log_sheet(wb):
    ws = wb.active
    ws.title = "Troubleshooting Log"
    ws.append(LOG_HEADERS)
    _style_header(ws, len(LOG_HEADERS))
    base = datetime(2026, 8, 10, 8, 30)
    for i, (product, issue, symptom, diag, res, status, mins, esc) in enumerate(CASES, start=1):
        when = base + timedelta(hours=i * 5)
        ws.append([
            f"CASE-{300 + i}", when.strftime("%Y-%m-%d %H:%M"), product, issue,
            symptom, diag, res, status, mins, esc,
        ])
    for idx, w in enumerate([10, 16, 14, 26, 26, 26, 30, 10, 16, 14], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    return ws


def build_kb_sheet(wb):
    ws = wb.create_sheet("KB Resolutions")
    ws.append(KB_HEADERS)
    _style_header(ws, len(KB_HEADERS))
    for row in KB:
        ws.append(list(row))
    for idx, w in enumerate([16, 34, 52, 20], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    return ws


def build_trend_sheet(wb, n_cases):
    ws = wb.create_sheet("Weekly Trend")
    ws["A1"] = "Weekly Troubleshooting Trend"
    ws["A1"].font = Font(bold=True, size=14)
    last = n_cases + 1

    ws["A3"] = "Cases by Product"
    ws["A3"].font = Font(bold=True)
    products = sorted({c[0] for c in CASES})
    for offset, product in enumerate(products, start=4):
        ws.cell(row=offset, column=1, value=product)
        ws.cell(row=offset, column=2,
                value=f'=COUNTIF(\'Troubleshooting Log\'!C2:C{last},"{product}")')

    col_e_start = 3
    ws["D3"] = "Escalation Metrics"
    ws["D3"].font = Font(bold=True)
    metrics = [
        ("Total cases", f"=COUNTA('Troubleshooting Log'!A2:A{last})"),
        ("Escalated (Y)", f'=COUNTIF(\'Troubleshooting Log\'!J2:J{last},"Y")'),
        ("Resolved", f'=COUNTIF(\'Troubleshooting Log\'!H2:H{last},"Resolved")'),
        ("Escalation rate", f'=COUNTIF(\'Troubleshooting Log\'!J2:J{last},"Y")/'
                            f"COUNTA('Troubleshooting Log'!A2:A{last})"),
        ("Total time (mins)", f"=SUM('Troubleshooting Log'!I2:I{last})"),
    ]
    for offset, (label, formula) in enumerate(metrics, start=4):
        ws.cell(row=offset, column=4, value=label)
        cell = ws.cell(row=offset, column=5, value=formula)
        if label == "Escalation rate":
            cell.number_format = "0.0%"
    for col, w in (("A", 18), ("B", 10), ("D", 18), ("E", 12)):
        ws.column_dimensions[col].width = w
    return ws


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "troubleshooting_log.xlsx"
    wb = Workbook()
    build_log_sheet(wb)
    build_kb_sheet(wb)
    build_trend_sheet(wb, len(CASES))
    wb.save(out)
    print(f"Wrote {out} with sheets: Troubleshooting Log, KB Resolutions, Weekly Trend "
          f"({len(CASES)} cases, {len(KB)} KB entries)")


if __name__ == "__main__":
    main()
